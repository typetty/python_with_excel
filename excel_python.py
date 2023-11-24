# This is the .py file for controlling and editing the excel files in this repository
# Contributers: Ty Petty
# Created on 11/17/23

'''
This application is to import the old excel file keeping track of student information
and use it to save all of the information on a new excel file in a more accessible 
and easy to read way.
'''


#import the necessary libraries
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import Font

# Find the Active Directory on any computer so that it is easy to save and find the external Excel files
activeDirectory = os.path.dirname(os.path.abspath(__file__))

# Open the external workbook that the information is already on
externalWorkbook = openpyxl.load_workbook(os.path.join(activeDirectory, "Poorly_Organized_Data_2.xlsx")) # Use the active directory to ensure it's found on every computer

# Set up the current sheet in the externalWorkbook to activeExternal
activeExternal = externalWorkbook.active

# Create a new workbook to reorganize information and to save at the end
newWorkbook = Workbook()
newWorkbook.remove(newWorkbook["Sheet"]) # Remove the blank sheet that the workbook autogenerates



# 1. Create new worksheets for each class
# Now create a for loop to parse through the data and make a different sheet for each class
for row_values in activeExternal.iter_rows(values_only=True, min_row=2):
    if row_values[0] not in newWorkbook.sheetnames:
        newWorkbook.create_sheet(row_values[0])


# 2. In each sheet, create columns for last name, first name, student ID, and grade with the student data for that class placed there.
for worksheet in newWorkbook:
    worksheet.append(["Last Name","First Name","Student ID","Grade"])
    # Now add in all the data below the titles of columns
    for row_values in activeExternal.iter_rows(values_only=True, min_row=2):
        if row_values[0] == worksheet.title:
            worksheet.append([row_values[1].split("_")[0],row_values[1].split("_")[1],row_values[1].split("_")[2],row_values[2]])


# 3. A filter should be placed over the 4 aforementioned columns in each sheet
for worksheet in newWorkbook:
    # Get the max row to  make the filter on the correct data
    maxRow = worksheet.max_row
    worksheet.auto_filter.ref = "A1:D" + str(maxRow)


# 4. each sheet should have some simple summary information about each class 
for worksheet in newWorkbook:
    #Will need the max row also for this
    maxRow = worksheet.max_row
    #insert all of the data
    worksheet["F1"] = "Summary Statistics"
    worksheet["G1"] = "Value"
    worksheet["F2"] = "Highest Grade"
    worksheet["G2"] = f'=MAX(D2:D{maxRow})'
    worksheet["F3"] = "Lowest Grade"
    worksheet["G3"] = f'=MIN(D2:D{maxRow})'
    worksheet["F4"] = "Mean Grade"
    worksheet["G4"] = f'=AVERAGE(D2:D{maxRow})'
    worksheet["F5"] = "Median Grade"
    worksheet["G5"] = f'=MEDIAN(D2:D{maxRow})'
    worksheet["F6"] = "Number of Students"
    worksheet["G6"] = f'=COUNT(D2:D{maxRow})'


# 5. 5.	Some simple formatting (bolding headers) and changing the width of the columns.
# First, bolding the headers
for worksheet in newWorkbook:
    for cell in worksheet[1]:
        cell.font = Font(bold=True) # Bold the whole first row
        newWidth = len(str(cell.value)) + 5 # Set the new width wanted to 5 more than the length of the header
        worksheet.column_dimensions[cell.column_letter].width = newWidth


# 6. Save the results as a new Excel file named “formatted_grades.xlsx”
# Save the new workbook in this same folder
newWorkbook.save(os.path.join(activeDirectory, "formatted_grades.xlsx"))
# Close the workbooks you have been accessing
externalWorkbook.close()
newWorkbook.close()
